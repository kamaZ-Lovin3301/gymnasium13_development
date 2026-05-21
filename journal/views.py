from datetime import datetime, timedelta, date
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.db.models import Avg
from openpyxl import Workbook
from .models import Teacher, Schedule, Class, Subject, Student, Grade, Homework, Parent, ParentStudent, Announcement
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
def get_teacher_from_request(request):
    #Возвращает объект Teacher
    if not request.user.is_authenticated:
        return None
    try:
        teacher = Teacher.objects.get(login=request.user.username)
        return teacher
    except Teacher.DoesNotExist:
        return None

def get_student_from_request(request):
    #Возвращает объект Student
    if not request.user.is_authenticated:
        return None
    try:
        student = Student.objects.get(login=request.user.username)
        return student
    except Student.DoesNotExist:
        return None

def get_default_teacher():
    return Teacher.objects.first()


def teacher_journal(request):
    teacher = get_teacher_from_request(request)
    if not teacher:
        return render(request, 'journal/no_teacher.html')
    classes_taught = Class.objects.filter(schedule__teacher=teacher).distinct()
    context = {
        'teacher': teacher,
        'classes': classes_taught,
    }
    return render(request, 'journal/select_class.html', context)


def select_subject(request, class_id):
    teacher = get_teacher_from_request(request)
    if not teacher:
        return redirect('teacher_journal')
    class_obj = get_object_or_404(Class, id=class_id)
    subjects = Subject.objects.filter(
        schedule__teacher=teacher,
        schedule__class_id=class_obj
    ).distinct()
    context = {
        'teacher': teacher,
        'class_obj': class_obj,
        'subjects': subjects,
    }
    return render(request, 'journal/select_subject.html', context)


#Страница журнала (оценки и ДЗ)
def journal_page(request, class_id, subject_id):
    teacher = get_teacher_from_request(request)
    if not teacher:
        return redirect('teacher_journal')
    class_obj = get_object_or_404(Class, id=class_id)
    subject_obj = get_object_or_404(Subject, id=subject_id)
    students = Student.objects.filter(class_id=class_obj)
    schedule_items = Schedule.objects.filter(
        class_id=class_obj,
        subject_id=subject_obj,
        teacher_id=teacher
    ).order_by('-day_of_week', '-lesson_number')
    selected_schedule = schedule_items.first()
    grades = {}
    if selected_schedule:
        grade_objects = Grade.objects.filter(
            schedule_id=selected_schedule,
            grade_date=date.today()
        )
        for g in grade_objects:
            grades[g.student_id] = g.grade
    homework = None
    if selected_schedule:
        homework = Homework.objects.filter(schedule=selected_schedule).first()
    context = {
        'teacher': teacher,
        'class_obj': class_obj,
        'subject_obj': subject_obj,
        'students': students,
        'schedule_items': schedule_items,
        'selected_schedule': selected_schedule,
        'grades': grades,
        'homework': homework,
        'today': date.today(),
    }
    return render(request, 'journal/journal_page.html', context)


#Выставить оценку
def set_grade(request):
    if request.method == 'POST':
        teacher = get_teacher_from_request(request)
        if not teacher:
            return JsonResponse({'error': 'Учитель не найден'}, status=403)

        student_id = request.POST.get('student_id')
        schedule_id = request.POST.get('schedule_id')
        grade_value = request.POST.get('grade')
        grade_date = request.POST.get('grade_date')

        if not all([student_id, schedule_id, grade_date]):
            return JsonResponse({'error': 'Не все данные переданы'}, status=400)

        schedule = get_object_or_404(Schedule, id=schedule_id)

        if schedule.teacher_id != teacher.id:
            return JsonResponse({'error': 'Нет прав'}, status=403)

        grade, created = Grade.objects.update_or_create(
            student_id=student_id,
            schedule=schedule,
            grade_date=grade_date,
            defaults={'grade': grade_value if grade_value else None}
        )

        return JsonResponse({'success': True, 'grade': grade_value})

    return JsonResponse({'error': 'Метод не разрешен'}, status=405)


#Сохранить домашнее задание
def set_homework(request):
    if request.method == 'POST':
        teacher = get_teacher_from_request(request)
        if not teacher:
            return JsonResponse({'error': 'Учитель не найден'}, status=403)

        schedule_id = request.POST.get('schedule_id')
        description = request.POST.get('description')
        file = request.FILES.get('file')

        if not schedule_id:
            return JsonResponse({'error': 'Не указан урок'}, status=400)

        schedule = get_object_or_404(Schedule, id=schedule_id)

        if schedule.teacher_id != teacher.id:
            return JsonResponse({'error': 'Нет прав'}, status=403)

        homework, created = Homework.objects.update_or_create(
            schedule=schedule,
            defaults={'description': description or ''}
        )

        if file:
            if homework.file:
                default_storage.delete(homework.file.name)
            file_name = f'homework/{schedule.id}_{file.name}'
            saved_path = default_storage.save(file_name, ContentFile(file.read()))
            homework.file = saved_path
            homework.save()

        return JsonResponse({
            'success': True,
            'file_url': homework.file.url if homework.file else None,
            'file_name': homework.file.name if homework.file else None
        })

    return JsonResponse({'error': 'Метод не разрешен'}, status=405)


def announcements_for_class(request):
    teacher = get_teacher_from_request(request)
    if not teacher or not teacher.homeroom_class:
        return redirect('teacher_journal')

    class_obj = teacher.homeroom_class

    if request.method == 'POST':
        title = request.POST.get('title')
        text = request.POST.get('text')
        if title and text:
            Announcement.objects.create(
                title=title,
                text=text,
                class_id=class_obj,
                teacher=teacher
            )
            return redirect('announcements_for_class')

    announcements = Announcement.objects.filter(class_id=class_obj).order_by('-created_at')
    return render(request, 'journal/announcements_for_class.html', {
        'teacher': teacher,
        'class_obj': class_obj,
        'announcements': announcements,
    })


def student_announcements(request):
    student = get_student_from_request(request)
    if not student:
        return redirect('teacher_journal')

    announcements = Announcement.objects.filter(class_id=student.class_id).order_by('-created_at')
    return render(request, 'journal/student_announcements.html', {
        'student': student,
        'announcements': announcements,
    })


def parent_announcements(request, child_id):
    parent = get_parent_from_request(request)
    if not parent:
        return redirect('teacher_journal')

    try:
        student = Student.objects.get(id=child_id, parentstudent__parent=parent)
    except Student.DoesNotExist:
        return redirect('parent_select_child')

    announcements = Announcement.objects.filter(class_id=student.class_id).order_by('-created_at')
    children = Student.objects.filter(parentstudent__parent=parent)

    return render(request, 'journal/parent_announcements.html', {
        'student': student,
        'announcements': announcements,
        'children': children,
        'parent': parent,
    })

def teacher_schedule(request):
    teacher = get_teacher_from_request(request)
    if not teacher:
        return redirect('teacher_journal')
    schedule = Schedule.objects.filter(teacher=teacher).select_related('class_id', 'subject', 'room').order_by(
        'day_of_week', 'lesson_number')
    days_map = {1: 'Понедельник', 2: 'Вторник', 3: 'Среда', 4: 'Четверг', 5: 'Пятница', 6: 'Суббота'}
    schedule_list = []
    for day in range(1, 7):
        day_lessons = [lesson for lesson in schedule if lesson.day_of_week == day]
        schedule_list.append({
            'day_name': days_map[day],
            'day_short': days_map[day][:2],
            'lessons': day_lessons
        })
    return render(request, 'journal/teacher_schedule.html', {
        'schedule_list': schedule_list,
        'teacher': teacher,
        'show_homeroom_link': teacher.homeroom_class is not None,
    })
def student_schedule(request):
    student = get_student_from_request(request)
    if not student:
        return redirect('teacher_journal')
    schedule = Schedule.objects.filter(class_id=student.class_id).select_related('subject', 'teacher', 'room').order_by(
        'day_of_week', 'lesson_number')

    days_map = {
        1: {'name': 'Понедельник', 'short': 'ПН'},
        2: {'name': 'Вторник', 'short': 'ВТ'},
        3: {'name': 'Среда', 'short': 'СР'},
        4: {'name': 'Четверг', 'short': 'ЧТ'},
        5: {'name': 'Пятница', 'short': 'ПТ'},
        6: {'name': 'Суббота', 'short': 'СБ'}
    }

    today = datetime.now().date()

    schedule_list = []
    for day in range(1, 7):
        day_lessons = []
        for lesson in schedule:
            if lesson.day_of_week == day:
                grade = Grade.objects.filter(
                    student=student,
                    schedule=lesson,
                    grade_date=today
                ).first()

                homework = Homework.objects.filter(schedule=lesson).first()

                day_lessons.append({
                    'lesson_number': lesson.lesson_number,
                    'subject': lesson.subject,
                    'teacher': lesson.teacher,
                    'room': lesson.room,
                    'grade': grade.grade if grade else None,
                    'homework': homework
                })

        day_lessons.sort(key=lambda x: x['lesson_number'])

        schedule_list.append({
            'day_name': days_map[day]['name'],
            'day_short': days_map[day]['short'],
            'lessons': day_lessons
        })

    current_month = today.strftime('%B %Y').capitalize()

    return render(request, 'journal/student_schedule.html', {
        'schedule_list': schedule_list,
        'student': student,
        'current_month': current_month
    })

def home(request):
    if request.user.is_authenticated:
        return redirect('teacher_journal')
    return redirect('login')
def redirect_after_login(request):
    if request.user.is_authenticated:
        try:
            parent = Parent.objects.get(login=request.user.username)
            return redirect('parent_select_child')
        except Parent.DoesNotExist:
            try:
                teacher = Teacher.objects.get(login=request.user.username)
                return redirect('teacher_schedule')
            except Teacher.DoesNotExist:
                try:
                    student = Student.objects.get(login=request.user.username)
                    return redirect('student_schedule')
                except Student.DoesNotExist:
                    return redirect('teacher_journal')
    return redirect('login')


def student_diary(request):
    student = get_student_from_request(request)
    if not student:
        return redirect('teacher_journal')

    date_str = request.GET.get('date')
    if date_str:
        current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        current_date = datetime.now().date()

    current_date_string = current_date.isoformat()

    start_of_month = current_date.replace(day=1)
    if current_date.month == 12:
        end_of_month = current_date.replace(year=current_date.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end_of_month = current_date.replace(month=current_date.month + 1, day=1) - timedelta(days=1)

    dates_range = []
    for d in range((end_of_month - start_of_month).days + 1):
        dates_range.append(start_of_month + timedelta(days=d))

    schedule_for_day = Schedule.objects.filter(
        class_id=student.class_id,
        day_of_week=current_date.isoweekday()
    ).select_related('subject', 'teacher', 'room')

    daily_schedule = []
    for schedule in schedule_for_day:
        grade = Grade.objects.filter(
            student=student,
            schedule=schedule,
            grade_date=current_date
        ).first()

        homework = Homework.objects.filter(
            schedule=schedule
        ).first()

        daily_schedule.append({
            'number': schedule.lesson_number,
            'subject': schedule.subject,
            'teacher': schedule.teacher,
            'room': schedule.room,
            'grade': grade.grade if grade else None,
            'homework': homework
        })

    daily_schedule.sort(key=lambda x: x['number'])

    return render(request, 'journal/student_diary.html', {
        'student': student,
        'daily_schedule': daily_schedule,
        'current_date': current_date,
        'current_date_string': current_date_string,
        'dates_range': dates_range
    })


def get_quarter_date_range(quarter, year=None):
    if year is None:
        year = date.today().year

    quarters = {
        1: {
            'start': date(year, 9, 1),
            'end': date(year, 10, 31)
        },
        2: {
            'start': date(year, 11, 1),
            'end': date(year, 12, 31)
        },
        3: {
            'start': date(year, 1, 1),
            'end': date(year, 3, 31)
        },
        4: {
            'start': date(year, 4, 1),
            'end': date(year, 5, 31)
        },
    }

    if quarter in [3, 4]:
        quarters[3]['start'] = date(year + 1, 1, 1)
        quarters[3]['end'] = date(year + 1, 3, 31)
        quarters[4]['start'] = date(year + 1, 4, 1)
        quarters[4]['end'] = date(year + 1, 5, 31)

    return quarters.get(quarter)


def quarter_grades(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    current_quarter = int(request.GET.get('quarter', 1))

    quarters = {
        1: {'months': [9, 10, 11]},
        2: {'months': [12, 1, 2]},
        3: {'months': [3, 4, 5]},
        4: {'months': [6, 7, 8]}
    }

    subjects = Subject.objects.filter(schedule__class_id=student.class_id).distinct()

    subjects_data = []
    for subject in subjects:
        grades = Grade.objects.filter(
            student=student,
            schedule__subject=subject,
            grade_date__month__in=quarters[current_quarter]['months']
        )

        grades_list = []
        for grade in grades:
            if grade.grade:
                grades_list.append(grade.grade)

        average = None
        if grades_list:
            avg = sum(grades_list) / len(grades_list)
            average = round(avg, 2)

        final_grade = None

        subjects_data.append({
            'name': subject.name,
            'grades': grades_list,
            'average': average,
            'final_grade': final_grade
        })

    context = {
        'student': student,
        'current_quarter': current_quarter,
        'subjects_data': subjects_data
    }

    return render(request, 'journal/quarter_grades.html', context)


def get_parent_from_request(request):
    if not request.user.is_authenticated:
        return None
    try:
        return Parent.objects.get(login=request.user.username)
    except Parent.DoesNotExist:
        return None


def parent_select_child(request):
    parent = get_parent_from_request(request)
    if not parent:
        return redirect('teacher_journal')

    children = Student.objects.filter(parentstudent__parent=parent)

    return render(request, 'journal/parent_select_child.html', {
        'parent': parent,
        'children': children
    })

def parent_schedule(request, child_id):
    parent = get_parent_from_request(request)
    if not parent:
        return redirect('teacher_journal')

    try:
        student = Student.objects.get(id=child_id, parentstudent__parent=parent)
    except Student.DoesNotExist:
        return redirect('parent_select_child')

    children = Student.objects.filter(parentstudent__parent=parent)

    schedule = Schedule.objects.filter(class_id=student.class_id).select_related('subject', 'teacher', 'room').order_by(
        'day_of_week', 'lesson_number')

    days_map = {1: 'ПН', 2: 'ВТ', 3: 'СР', 4: 'ЧТ', 5: 'ПТ', 6: 'СБ'}
    today = datetime.now().date()

    schedule_list = []
    for day in range(1, 7):
        day_lessons = []
        for lesson in schedule:
            if lesson.day_of_week == day:
                grade = Grade.objects.filter(student=student, schedule=lesson, grade_date=today).first()
                homework = Homework.objects.filter(schedule=lesson).first()
                day_lessons.append({
                    'lesson_number': lesson.lesson_number,
                    'subject': lesson.subject,
                    'teacher': lesson.teacher,
                    'room': lesson.room,
                    'grade': grade.grade if grade else None,
                    'homework': homework
                })
        day_lessons.sort(key=lambda x: x['lesson_number'])
        schedule_list.append({'day_short': days_map[day], 'lessons': day_lessons})

    return render(request, 'journal/parent_schedule.html', {
        'schedule_list': schedule_list,
        'student': student,
        'parent': parent,
        'children': children,
        'current_month': today.strftime('%B %Y').capitalize()
    })

def parent_diary(request, child_id):
    parent = get_parent_from_request(request)
    if not parent:
        return redirect('teacher_journal')

    try:
        student = Student.objects.get(id=child_id, parentstudent__parent=parent)
    except Student.DoesNotExist:
        return redirect('parent_select_child')

    children = Student.objects.filter(parentstudent__parent=parent)

    date_str = request.GET.get('date')
    if date_str:
        current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        current_date = datetime.now().date()

    current_date_string = current_date.isoformat()

    start_of_month = current_date.replace(day=1)
    if current_date.month == 12:
        end_of_month = current_date.replace(year=current_date.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end_of_month = current_date.replace(month=current_date.month + 1, day=1) - timedelta(days=1)

    dates_range = []
    for d in range((end_of_month - start_of_month).days + 1):
        dates_range.append(start_of_month + timedelta(days=d))

    schedule_for_day = Schedule.objects.filter(
        class_id=student.class_id,
        day_of_week=current_date.isoweekday()
    ).select_related('subject', 'teacher', 'room')

    daily_schedule = []
    for schedule in schedule_for_day:
        grade = Grade.objects.filter(student=student, schedule=schedule, grade_date=current_date).first()
        homework = Homework.objects.filter(schedule=schedule).first()
        daily_schedule.append({
            'number': schedule.lesson_number,
            'subject': schedule.subject,
            'teacher': schedule.teacher,
            'room': schedule.room,
            'grade': grade.grade if grade else None,
            'homework': homework
        })

    daily_schedule.sort(key=lambda x: x['number'])

    return render(request, 'journal/parent_diary.html', {
        'student': student,
        'daily_schedule': daily_schedule,
        'current_date': current_date,
        'current_date_string': current_date_string,
        'dates_range': dates_range,
        'parent': parent,
        'children': children,
    })

def parent_quarter_grades(request, child_id):
    parent = get_parent_from_request(request)
    if not parent:
        return redirect('teacher_journal')

    try:
        student = Student.objects.get(id=child_id, parentstudent__parent=parent)
    except Student.DoesNotExist:
        return redirect('parent_select_child')

    children = Student.objects.filter(parentstudent__parent=parent)

    current_quarter = int(request.GET.get('quarter', 1))

    quarters = {
        1: {'months': [9, 10, 11]},
        2: {'months': [12, 1, 2]},
        3: {'months': [3, 4, 5]},
        4: {'months': [6, 7, 8]}
    }

    subjects = Subject.objects.filter(schedule__class_id=student.class_id).distinct()

    subjects_data = []
    for subject in subjects:
        grades = Grade.objects.filter(
            student=student,
            schedule__subject=subject,
            grade_date__month__in=quarters[current_quarter]['months']
        )
        grades_list = [g.grade for g in grades if g.grade]
        average = round(sum(grades_list) / len(grades_list), 2) if grades_list else None
        subjects_data.append({
            'name': subject.name,
            'grades': grades_list,
            'average': average,
            'final_grade': None
        })

    return render(request, 'journal/parent_quarter_grades.html', {
        'student': student,
        'current_quarter': current_quarter,
        'subjects_data': subjects_data,
        'parent': parent,
        'children': children,
    })

from django.http import HttpResponse
from openpyxl import Workbook

def homeroom_dashboard(request):
    teacher = get_teacher_from_request(request)
    if not teacher:
        return redirect('teacher_journal')

    homeroom_class = teacher.homeroom_class
    if not homeroom_class:
        return render(request, 'journal/not_homeroom.html', {'teacher': teacher})

    students = Student.objects.filter(class_id=homeroom_class).order_by('last_name', 'first_name')

    subjects = Subject.objects.filter(schedule__class_id=homeroom_class).distinct().order_by('name')

    selected_subject_id = request.GET.get('subject_id')
    selected_subject = None
    students_data = []

    if selected_subject_id:
        selected_subject = get_object_or_404(Subject, id=selected_subject_id)
        for student in students:
            grades = Grade.objects.filter(
                student=student,
                schedule__subject=selected_subject,
                schedule__class_id=homeroom_class
            ).order_by('grade_date')  # сортируем по дате
            grade_values = [g.grade for g in grades if g.grade is not None]
            avg_grade = round(sum(grade_values) / len(grade_values), 2) if grade_values else None
            students_data.append({
                'student': student,
                'grades': grades,
                'avg_grade': avg_grade,
            })

    context = {
        'teacher': teacher,
        'homeroom_class': homeroom_class,
        'students': students,
        'subjects': subjects,
        'selected_subject': selected_subject,
        'students_data': students_data,
    }
    return render(request, 'journal/homeroom_dashboard.html', context)


def export_homeroom_grades(request):
    teacher = get_teacher_from_request(request)
    if not teacher or not teacher.homeroom_class:
        return redirect('teacher_journal')

    homeroom_class = teacher.homeroom_class
    subject_id = request.GET.get('subject_id')
    if not subject_id:
        return redirect('homeroom_dashboard')

    subject = get_object_or_404(Subject, id=subject_id)
    students = Student.objects.filter(class_id=homeroom_class).order_by('last_name', 'first_name')

    wb = Workbook()
    ws = wb.active
    ws.title = f"{homeroom_class.grade}{homeroom_class.letter}_{subject.name}"

    ws['A1'] = 'Фамилия'
    ws['B1'] = 'Имя'
    ws['C1'] = 'Средний балл'
    ws['D1'] = 'Оценки (через запятую)'

    row = 2
    for student in students:
        grades = Grade.objects.filter(
            student=student,
            schedule__subject=subject,
            schedule__class_id=homeroom_class
        )
        grade_values = [str(g.grade) for g in grades if g.grade is not None]
        avg_grade = round(sum(map(int, grade_values)) / len(grade_values), 2) if grade_values else ''
        ws.cell(row=row, column=1, value=student.last_name)
        ws.cell(row=row, column=2, value=student.first_name)
        ws.cell(row=row, column=3, value=avg_grade)
        ws.cell(row=row, column=4, value=', '.join(grade_values))
        row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="success_{homeroom_class.grade}{homeroom_class.letter}_{subject.name}.xlsx"'
    wb.save(response)
    return response

def guest_home(request):
    return render(request, 'journal/guest_home.html')
